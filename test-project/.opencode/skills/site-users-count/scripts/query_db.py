#!/usr/bin/env python3
"""
Query online_users_count table from the wireless database (PostgreSQL).
Read-only - only SELECT queries are allowed.

Usage:
  python scripts/query_db.py --site-code HQ --date 2024-01-15
  python scripts/query_db.py --site-name "Beijing" --start-date 2024-01-01 --end-date 2024-01-31 --output excel
  python scripts/query_db.py --site-code HQ --start-date 2024-01-01 --end-date 2024-01-31 --output chart --chart-type line
  python scripts/query_db.py --sql "SELECT site_name, AVG(count) as avg_count FROM online_users_count GROUP BY site_name" --output json
  python scripts/query_db.py --list-sites
  python scripts/query_db.py --group-by site_owner --start-date 2024-01-01 --end-date 2024-01-31 --output chart --chart-type bar
"""

import argparse
import json
import os
import sys
import re
from datetime import datetime, timedelta, date
from decimal import Decimal
from pathlib import Path

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

FORBIDDEN_KEYWORDS = re.compile(
    r'\b(INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|TRUNCATE|REPLACE|RENAME|GRANT|REVOKE|LOCK|UNLOCK|CALL|EXEC|EXECUTE|INTO\s+OUTFILE|LOAD\s+DATA|COPY|VACUUM|REINDEX|CLUSTER)\b',
    re.IGNORECASE
)


def get_connection(env_path=None):
    if env_path is None:
        for candidate in [Path.cwd() / '.env', Path.cwd() / '.env.local']:
            if candidate.exists():
                env_path = candidate
                break
    if env_path and Path(env_path).exists():
        load_dotenv(env_path)
    else:
        load_dotenv()

    host = os.getenv('DB_HOST', 'localhost')
    port = int(os.getenv('DB_PORT', '5432'))
    user = os.getenv('DB_USER', 'postgres')
    password = os.getenv('DB_PASSWORD', '')
    database = os.getenv('DB_NAME', 'wireless')

    conn = psycopg2.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        dbname=database,
        cursor_factory=psycopg2.extras.RealDictCursor,
    )
    conn.set_session(readonly=True, autocommit=True)
    return conn


def validate_sql(sql):
    if FORBIDDEN_KEYWORDS.search(sql):
        raise ValueError("Only SELECT queries are allowed. Detected forbidden SQL keywords.")
    sql_upper = sql.strip().upper()
    if not sql_upper.startswith('SELECT') and not sql_upper.startswith('SHOW') and not sql_upper.startswith('DESCRIBE') and not sql_upper.startswith('EXPLAIN'):
        raise ValueError("Only SELECT/SHOW/DESCRIBE/EXPLAIN queries are allowed.")
    return True


def execute_query(conn, sql, params=None):
    validate_sql(sql)
    with conn.cursor() as cursor:
        cursor.execute(sql, params)
        results = cursor.fetchall()
    return [dict(row) for row in results]


def query_by_site_and_date(conn, site_code=None, site_name=None, date=None,
                           start_date=None, end_date=None, extra_where=None):
    conditions = []
    params = []

    if site_code:
        conditions.append("site_code = %s")
        params.append(site_code)
    if site_name:
        conditions.append("site_name LIKE %s")
        params.append(f"%{site_name}%")
    if date:
        conditions.append("create_time::date = %s")
        params.append(date)
    if start_date:
        conditions.append("create_time::date >= %s")
        params.append(start_date)
    if end_date:
        conditions.append("create_time::date <= %s")
        params.append(end_date)
    if extra_where:
        conditions.append(extra_where)

    where_clause = " AND ".join(conditions) if conditions else "1=1"
    sql = f"""
        SELECT id, site_code, site_name, count, count_nac, count_controller,
               site_owner, site_function, priority, business,
               create_time::date as record_date, create_time, update_time
        FROM online_users_count
        WHERE {where_clause}
        ORDER BY create_time DESC
    """
    return execute_query(conn, sql, params if params else None)


def list_sites(conn):
    sql = """
        SELECT DISTINCT site_code, site_name, site_owner, site_function, priority, business
        FROM online_users_count
        ORDER BY site_code
    """
    return execute_query(conn, sql)


def aggregate_query(conn, group_by, start_date=None, end_date=None, agg_func='SUM', agg_field='count'):
    allowed_group = ['site_code', 'site_name', 'site_owner', 'site_function', 'priority', 'business',
                     'create_time::date', 'EXTRACT(YEAR FROM create_time)',
                     'EXTRACT(MONTH FROM create_time)',
                     'TO_CHAR(create_time, \'IYYY-IW\')']
    if group_by not in allowed_group:
        raise ValueError(f"group_by must be one of: {allowed_group}")

    conditions = []
    params = []
    if start_date:
        conditions.append("create_time::date >= %s")
        params.append(start_date)
    if end_date:
        conditions.append("create_time::date <= %s")
        params.append(end_date)

    where_clause = " AND ".join(conditions) if conditions else "1=1"

    group_map = {
        'create_time::date': ("create_time::date", "create_time::date as group_key"),
        'EXTRACT(YEAR FROM create_time)': ("EXTRACT(YEAR FROM create_time)", "EXTRACT(YEAR FROM create_time)::int as group_key"),
        'EXTRACT(MONTH FROM create_time)': ("EXTRACT(MONTH FROM create_time)", "EXTRACT(MONTH FROM create_time)::int as group_key"),
        "TO_CHAR(create_time, 'IYYY-IW')": ("TO_CHAR(create_time, 'IYYY-IW')", "TO_CHAR(create_time, 'IYYY-IW') as group_key"),
    }

    if group_by in group_map:
        group_col, select_col = group_map[group_by]
    else:
        select_col = f'{group_by} as group_key'
        group_col = group_by

    sql = f"""
        SELECT {select_col},
               {agg_func}({agg_field}) as agg_value,
               COUNT(*) as record_count,
               AVG({agg_field}) as avg_value,
               MIN({agg_field}) as min_value,
               MAX({agg_field}) as max_value
        FROM online_users_count
        WHERE {where_clause}
        GROUP BY {group_col}
        ORDER BY {group_col}
    """
    return execute_query(conn, sql, params if params else None)


def compare_sites(conn, site_codes, start_date=None, end_date=None):
    if not site_codes:
        raise ValueError("At least one site_code is required for comparison")

    conditions = []
    params = []

    placeholders = ', '.join(['%s'] * len(site_codes))
    conditions.append(f"site_code IN ({placeholders})")
    params.extend(site_codes)

    if start_date:
        conditions.append("create_time::date >= %s")
        params.append(start_date)
    if end_date:
        conditions.append("create_time::date <= %s")
        params.append(end_date)

    where_clause = " AND ".join(conditions)
    sql = f"""
        SELECT site_code, site_name,
               create_time::date as record_date,
               count, count_nac, count_controller
        FROM online_users_count
        WHERE {where_clause}
        ORDER BY site_code, create_time
    """
    return execute_query(conn, sql, params)


def to_excel(data, output_path, title=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title or "Online Users Count"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if not data:
        ws['A1'] = "No data found"
        wb.save(output_path)
        return output_path

    headers = list(data[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row.get(header)
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, timedelta):
                value = str(value)
            elif isinstance(value, Decimal):
                value = float(value)
            elif isinstance(value, date) and not isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in data:
            val = str(row.get(header, ''))
            if len(val) > max_length:
                max_length = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 50)

    wb.save(output_path)
    return output_path


def to_chart(data, output_path, x_field='group_key', y_field='agg_value',
             chart_type='bar', title='Online Users Count', label_field=None):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(14, 7))

    if chart_type == 'line':
        x_labels = [str(row.get(x_field, '')) for row in data]
        y_values = [float(row.get(y_field, 0) or 0) for row in data]
        ax.plot(x_labels, y_values, marker='o', linewidth=2, markersize=5)
        ax.fill_between(range(len(y_values)), y_values, alpha=0.1)
        ax.set_xticklabels(x_labels, rotation=45, ha='right')
    elif chart_type == 'bar':
        x_labels = [str(row.get(x_field, '')) for row in data]
        y_values = [float(row.get(y_field, 0) or 0) for row in data]
        ax.bar(x_labels, y_values, color='#4472C4')
        ax.set_xticklabels(x_labels, rotation=45, ha='right')
    elif chart_type == 'pie':
        labels = [str(row.get(x_field, '')) for row in data]
        values = [float(row.get(y_field, 0) or 0) for row in data]
        ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
        ax.axis('equal')
    elif chart_type == 'multi_line':
        if label_field:
            groups = {}
            for row in data:
                label = str(row.get(label_field, ''))
                x_val = str(row.get(x_field, ''))
                y_val = float(row.get(y_field, 0) or 0)
                if label not in groups:
                    groups[label] = {'x': [], 'y': []}
                groups[label]['x'].append(x_val)
                groups[label]['y'].append(y_val)

            for label, group_data in groups.items():
                ax.plot(group_data['x'], group_data['y'], marker='o', linewidth=2, label=label, markersize=4)

            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
            x_labels_set = set()
            for g in groups.values():
                x_labels_set.update(g['x'])
            ax.set_xticklabels(sorted(x_labels_set), rotation=45, ha='right')
        else:
            x_labels = [str(row.get(x_field, '')) for row in data]
            y_values = [float(row.get(y_field, 0) or 0) for row in data]
            ax.plot(x_labels, y_values, marker='o', linewidth=2)
            ax.set_xticklabels(x_labels, rotation=45, ha='right')

    if chart_type != 'pie':
        ax.set_xlabel(x_field, fontsize=12)
        ax.set_ylabel(y_field, fontsize=12)

    ax.set_title(title, fontsize=14, fontweight='bold')
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return output_path


def format_as_table(data):
    if not data:
        return "No data found."

    headers = list(data[0].keys())
    col_widths = []
    for h in headers:
        max_w = len(str(h))
        for row in data:
            val = str(row.get(h, ''))
            if len(val) > max_w:
                max_w = len(val)
        col_widths.append(min(max_w, 40))

    header_line = " | ".join(str(h).ljust(col_widths[i]) for i, h in enumerate(headers))
    separator = "-+-".join("-" * w for w in col_widths)

    lines = [header_line, separator]
    for row in data:
        line = " | ".join(str(row.get(h, ''))[:col_widths[i]].ljust(col_widths[i]) for i, h in enumerate(headers))
        lines.append(line)

    return "\n".join(lines)


def serialize_row(row):
    s = {}
    for k, v in row.items():
        if isinstance(v, datetime):
            s[k] = v.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(v, date) and not isinstance(v, datetime):
            s[k] = v.strftime('%Y-%m-%d')
        elif isinstance(v, Decimal):
            s[k] = float(v)
        elif isinstance(v, timedelta):
            s[k] = str(v)
        else:
            s[k] = v
    return s


def main():
    parser = argparse.ArgumentParser(description='Query online_users_count from wireless database (read-only)')
    parser.add_argument('--env-path', help='Path to .env file')
    parser.add_argument('--site-code', help='Filter by site_code')
    parser.add_argument('--site-name', help='Filter by site_name (fuzzy match)')
    parser.add_argument('--date', help='Specific date (YYYY-MM-DD)')
    parser.add_argument('--start-date', help='Start date (YYYY-MM-DD)')
    parser.add_argument('--end-date', help='End date (YYYY-MM-DD)')
    parser.add_argument('--group-by', help='Group by field for aggregation')
    parser.add_argument('--agg-func', default='SUM', help='Aggregation function: SUM, AVG, MAX, MIN, COUNT')
    parser.add_argument('--agg-field', default='count', help='Field to aggregate: count, count_nac, count_controller')
    parser.add_argument('--compare-sites', nargs='+', help='Site codes to compare')
    parser.add_argument('--list-sites', action='store_true', help='List all distinct sites')
    parser.add_argument('--sql', help='Custom SELECT query (read-only)')
    parser.add_argument('--output', choices=['table', 'json', 'csv', 'excel', 'chart'], default='table')
    parser.add_argument('--output-path', help='Output file path (for excel/chart/csv)')
    parser.add_argument('--chart-type', choices=['line', 'bar', 'pie', 'multi_line'], default='bar')
    parser.add_argument('--chart-x', default='group_key', help='X-axis field for chart')
    parser.add_argument('--chart-y', default='agg_value', help='Y-axis field for chart')
    parser.add_argument('--chart-label', help='Label field for multi-line chart')
    parser.add_argument('--chart-title', help='Chart title')
    parser.add_argument('--limit', type=int, default=1000, help='Max rows to return (default 1000)')

    args = parser.parse_args()

    try:
        conn = get_connection(args.env_path)
        data = None

        if args.list_sites:
            data = list_sites(conn)
        elif args.sql:
            data = execute_query(conn, args.sql)
        elif args.compare_sites:
            data = compare_sites(conn, args.compare_sites, args.start_date, args.end_date)
        elif args.group_by:
            data = aggregate_query(conn, args.group_by, args.start_date, args.end_date,
                                   args.agg_func, args.agg_field)
        else:
            data = query_by_site_and_date(conn, args.site_code, args.site_name,
                                          args.date, args.start_date, args.end_date)

        if data and args.limit:
            data = data[:args.limit]

        if args.output == 'json':
            serialized = [serialize_row(row) for row in (data or [])]
            print(json.dumps(serialized, ensure_ascii=False, indent=2))

        elif args.output == 'csv':
            output_path = args.output_path or 'query_result.csv'
            if data:
                import csv
                headers = list(data[0].keys())
                with open(output_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=headers)
                    writer.writeheader()
                    for row in data:
                        writer.writerow(serialize_row(row))
                print(f"CSV saved to {output_path}")
            else:
                print("No data to export.")

        elif args.output == 'excel':
            output_path = args.output_path or 'online_users_report.xlsx'
            to_excel(data, output_path)
            print(f"Excel saved to {output_path}")

        elif args.output == 'chart':
            output_path = args.output_path or 'chart.png'
            to_chart(data, output_path,
                     x_field=args.chart_x,
                     y_field=args.chart_y,
                     chart_type=args.chart_type,
                     title=args.chart_title or 'Online Users Count',
                     label_field=args.chart_label)
            print(f"Chart saved to {output_path}")

        else:
            print(format_as_table(data))
            if data:
                print(f"\n({len(data)} rows)")

    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except psycopg2.OperationalError as e:
        print(f"Database connection error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        try:
            conn.close()
        except Exception:
            pass


if __name__ == '__main__':
    main()
